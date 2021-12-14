from openpyxl import load_workbook
import os
import json
from collections import OrderedDict
import ast

load_wb1 = load_workbook('./yoonho/best_PPT.xlsx')
load_ws1 = load_wb1['Sheet']
load_wb2 = load_workbook('./yoonho/APSC_c.xlsx')
load_ws2 = load_wb2['Sheet']

final_data = []
scenario_list = [[1], [1]*2, [1]*4, [1,3], [1,1,2], [2,2], [1]*8, [1,7], [2,6], [3,5], [4,4], [2,2,2,2], [1,1,3,3], [1]*100]

for i in range(27):
    dic = OrderedDict()
    name_to_id = OrderedDict()
    file_name = load_ws1.cell(1, i+2).value
    dic['filename'] = file_name

    with open('./data/' + file_name, 'r') as f:
        json_data = json.load(f)
    f.close()

    json_data = json_data["data"]

    for k in range(len(json_data)):
        test = json_data[k]
        test_name = test["name"]
        name_to_id[test_name] = k+1

    dic['name_to_id'] = name_to_id

    result = []
    for j in range(2, 16):
        answer = load_ws1.cell(j, i+2).value
        score = load_ws2.cell(j, i+2).value
        perm = ast.literal_eval(answer)
        
        result_dic = OrderedDict()
        result_dic['scenario'] = scenario_list[j-2]
        result_dic['TC'] = 1.5
        result_dic['permutations'] = perm
        result_dic['APSC_c'] = score
        result.append(result_dic)
    
    dic['results'] = result

    final_data.append(dic)

with open('./yoonho/genetic_result.json', 'w', encoding='utf-8') as make_file:
    json.dump(final_data, make_file, ensure_ascii=False, indent='\t')