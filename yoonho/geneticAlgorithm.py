import json
import random
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook


#parameters
GENERATION_LIMIT = 100
POPULATION_SIZE = 100
MUTATION_RATE = 0.1
TOURNAMENT_SELECTION = 10
PARENTS_NUM = 20 # number of parents which generate offspring by crossover


def average_time(resource_num, test_time_list, performance_list):
    sample = [[] for i in range(resource_num)]

    sorted_test_time_list = test_time_list[:]
    sorted_test_time_list.sort()
    sorted_test_time_list.reverse()

    for time in sorted_test_time_list:
        temp = list(map(lambda x:sum(x), sample))
        
        for i in range(len(temp)):
            temp[i] /= performance_list[i]
        
        idx = 0
        min_time = temp[0]

        for i in range(len(temp)):
            if temp[i] < min_time:
                min_time = temp[i]
                idx = i
            if temp[i] == min_time:
                if performance_list[i] > performance_list[idx]:
                    idx = i
        
        sample[idx].append(time)
    
    return max(list(map(lambda x:sum(x), sample)))


def test_seq_time(test_seq, test_time_list, performance_rate):
    time = 0.0
    
    for test_idx in test_seq:
        time += test_time_list[test_idx]/performance_rate

    return time


def individual_time(individual, test_time_list, performance_list):
    lst = []

    for i in range(len(individual)):
        test_seq = individual[i]
        lst.append(test_seq_time(test_seq, test_time_list, performance_list[i]))
    
    return max(lst)


def test_startTime_endTime_dict(individual, test_time_list, performance_list):
    dict1 = {}
    dict2 = {}

    for i in range(len(individual)):
        test_seq = individual[i]
        performance_rate = performance_list[i]
        time = 0.0
        for test_idx in test_seq:
            dict1[test_idx] = time
            time += test_time_list[test_idx]/performance_rate
            dict2[test_idx] = time


    return dict1, dict2


def satisfy_time_constraint(individual, time_constraint, test_time_list, performance_list):
    for i in range(len(individual)):
        test_seq = individual[i]
        time = test_seq_time(test_seq, test_time_list, performance_list[i])
        
        if time > time_constraint:
            return False
        
    return True


def shortest_test_seq(individual, test_time_list, performance_list):
    idx = 0
    min_time = 1000.0

    for i in range(len(individual)):
        test_seq = individual[i]
        time = test_seq_time(test_seq, test_time_list, performance_list[i])
        
        if time < min_time:
            min_time = time
            idx = i
        
        if time == min_time:
            if performance_list[i] > performance_list[idx]:
                idx = i
    
    return idx


def initialization(population_size, test_time_list, time_constarint, resource_num, performance_list):
    population = []

    while len(population) < population_size:
        permutation = [i for i in range(len(test_time_list))]
        random.shuffle(permutation)

        individual = [[] for _ in range(resource_num)]

        for test_idx in permutation:
            individual[shortest_test_seq(individual, test_time_list, performance_list)].append(test_idx)
        
        if satisfy_time_constraint(individual, time_constarint, test_time_list, performance_list):
            population.append(individual)

    return population


def find_index(individual, test_idx):
    for i in range(len(individual)):
        test_seq = individual[i]
        for j in range(len(test_seq)):
            if test_seq[j] == test_idx:
                return i, j
    print('error: no such index')


def fitness(individual, file_name_list, file_lines_list, test_time_list, test_coverage_list, performance_list):
    first_test_for_line = []
    for i in range(len(file_lines_list)):
        length = len(file_lines_list[i])
        first_test_for_line.append([-1 for i in range(length)])

    seq = []

    startTime_dict, _ = test_startTime_endTime_dict(individual, test_time_list, performance_list)
    while len(startTime_dict) != 0:
        idx = min(startTime_dict.keys(), key=lambda k: startTime_dict[k])
        seq.append(idx)
        del startTime_dict[idx]

    for test_idx in seq:
        assert test_idx >= 0
        test_coverage = test_coverage_list[test_idx]
            
        for content in test_coverage:
            file_name = content['class']
            lines = content['lines']

            file_idx = file_name_list.index(file_name)
            file_lines = file_lines_list[file_idx]

            for line in lines:
                line_idx = file_lines.index(line)
                if first_test_for_line[file_idx][line_idx] == -1:
                    first_test_for_line[file_idx][line_idx] = test_idx

    m = sum(list(map(len, first_test_for_line)))
    total_time = sum(test_time_list)
    answer = 0.0

    help_calculation = []
    for test_idx in range(len(test_time_list)):
        temp = total_time
        i, j = find_index(individual, test_idx)
        test_seq = individual[i]
        performance_rate = performance_list[i]
        for x in range(j+1):
            temp -= test_time_list[test_seq[x]]/performance_rate
        temp += 0.5 * test_time_list[test_seq[j]]/performance_rate
        help_calculation.append(temp)

    for item in first_test_for_line:
        for test_idx in item:
            assert test_idx >= 0, 'test_idx: {}'.format(test_idx)
            answer += help_calculation[test_idx]
            # if test_idx == -1:
            #     m -= 1
            # else:
            #     answer += help_calculation[test_idx]
    
    return round(answer/(total_time * m), 4)


def evaluate(population, file_name_list, file_lines_list, test_time_list, test_coverage_list, performance_list):
    evaluation = []

    for individual in population:
        evaluation.append(fitness(individual, file_name_list, file_lines_list, test_time_list, test_coverage_list, performance_list))

    return evaluation


def selection(population, evaluation):
    parents = []

    idx_list = list(range(len(population)))

    while len(parents) < PARENTS_NUM:
        selected_idx = random.sample(idx_list, TOURNAMENT_SELECTION)
        
        idx = 0
        max_fitness = 0.0

        for i in selected_idx:
            fitness = evaluation[i]
            
            if fitness > max_fitness:
                max_fitness = fitness
                idx = i
        
        parents.append(population[idx])
        idx_list.remove(idx)

    return parents


def crossover(parents, test_time_list, time_constraint, performance_list):
    population = parents

    while len(population) < POPULATION_SIZE:
        parent1, parent2 = random.sample(parents, 2)
        resource_num = len(parent1)

        time1 = individual_time(parent1, test_time_list, performance_list)
        time2 = individual_time(parent2, test_time_list, performance_list)
        time = max(time1, time2)

        cutpoint = random.random() * time

        parent1_startTime_dict, parent1_endTime_dict = test_startTime_endTime_dict(parent1, test_time_list, performance_list)
        parent2_startTime_dict, parent2_endTime_dict = test_startTime_endTime_dict(parent2, test_time_list, performance_list)

        offspring1 = [[] for _ in range(resource_num)]
        offspring2 = [[] for _ in range(resource_num)]

        for i in range(resource_num):
            test_seq = parent1[i]

            for test_idx in test_seq:
                if parent1_endTime_dict[test_idx] < cutpoint:
                    offspring1[i].append(test_idx)
                    del parent2_startTime_dict[test_idx]
                else:
                    break

        while len(parent2_startTime_dict) != 0:
            idx = min(parent2_startTime_dict.keys(), key=lambda k: parent2_startTime_dict[k])
            offspring1[shortest_test_seq(offspring1, test_time_list, performance_list)].append(idx)
            del parent2_startTime_dict[idx]

        if satisfy_time_constraint(offspring1, time_constraint, test_time_list, performance_list):
            population.append(offspring1)


        for i in range(resource_num):
            test_seq = parent2[i]

            for test_idx in test_seq:
                if parent2_endTime_dict[test_idx] < cutpoint:
                    offspring2[i].append(test_idx)
                    del parent1_startTime_dict[test_idx]
                else:
                    break

        while len(parent1_startTime_dict) != 0:
            idx = min(parent1_startTime_dict.keys(), key=lambda k: parent1_startTime_dict[k])
            offspring2[shortest_test_seq(offspring2, test_time_list, performance_list)].append(idx)
            del parent1_startTime_dict[idx]

        if satisfy_time_constraint(offspring2, time_constraint, test_time_list, performance_list):
            population.append(offspring2)

    if len(population) > POPULATION_SIZE:
        del population[-1]

    return population


def mutation(population, test_time_list, time_constraint, performance_list):
    for i in range(len(population)):
        rand = random.random()
        
        if rand < MUTATION_RATE:
            individual = population[i]
            
            temp_list = list(range(len(individual)))
            test_seq_idx = random.choice(temp_list)
            test_seq = individual[test_seq_idx]

            while len(test_seq) == 1:
                temp_list.remove(test_seq_idx)
                test_seq_idx = random.choice(temp_list)
                test_seq = individual[test_seq_idx]
            
            idx1, idx2 = random.sample(range(len(test_seq)), 2)

            test_seq[idx1], test_seq[idx2] = test_seq[idx2], test_seq[idx1]

    return   
            



def geneticAlgorithm(resource_num, performance_list, path):

    with open(path, 'r') as f:
        json_data = json.load(f)
    f.close()

    test_name_list = []
    test_time_list = []
    test_coverage_list = []

    json_data = json_data["data"]

    for i in range(len(json_data)):
        test = json_data[i]
        
        if 'coverage' in test:
            test_name_list.append(test['name'])
            if test['duration'] == 0.0:
                test_time_list.append(0.0001)
            else:
                test_time_list.append(test['duration'])
            test_coverage_list.append(test['coverage'])
    
    file_name_list = []
    file_lines_list = []

    for coverage in test_coverage_list:
        for file_and_line in coverage:
            file_name = file_and_line['class']
            file_lines = file_and_line['lines'] #list of line

            if file_name not in file_name_list:
                file_name_list.append(file_name)
                file_lines_list.append(file_lines)
            
            else:
                idx = file_name_list.index(file_name)
                file_lines_list[idx] = list(set(file_lines_list[idx] + file_lines))


    generation = 0
    graph1 = []
    graph2 = []

    average = average_time(resource_num, test_time_list, performance_list)
    print("average:", average)
    time_constraint = round(1.5*average, 4)

    population = initialization(POPULATION_SIZE, test_time_list, time_constraint, resource_num, performance_list)
    evaluation = evaluate(population, file_name_list, file_lines_list, test_time_list, test_coverage_list, performance_list)
    graph1.append(max(evaluation))
    # graph2.append(sum(evaluation)/len(evaluation))

    while generation < GENERATION_LIMIT:
        if generation % 10 == 0:
            print("generation{}...".format(generation))
        generation += 1
        # print('selection')
        parents = selection(population, evaluation)
        # print('crossover')
        population = crossover(parents, test_time_list, time_constraint, performance_list)
        # print('mutation')
        mutation(population, test_time_list, time_constraint, performance_list)
        # print('evaluation')
        evaluation = evaluate(population, file_name_list, file_lines_list, test_time_list, test_coverage_list, performance_list)
        graph1.append(max(evaluation))
        # graph2.append(sum(evaluation)/len(evaluation))
    
    best_fitness = max(evaluation)
    idx = evaluation.index(best_fitness)
    best_individual = population[idx]

    # plt.plot([i for i in range(GENERATION_LIMIT+1)], graph1)
    # plt.ylabel('fitness')
    # plt.xlabel('generation')
    # plt.show()

    for test_seq in best_individual:
        for i in range(len(test_seq)):
            test_idx = test_seq[i]
            test_seq[i] = test_name_list[test_idx]

    return best_individual, best_fitness


def main():
    resource = [1, 2, 4, 8, 16, 2, 2, 2, 8]
    performance = [[1], [1,1], [1,1,1,1], [1 for i in range(8)], [1 for i in range(16)], [1,2], [1,3], [1,4], [1,1,1,1,4,4,4,4]]

    dir_path = "./data"
    file_list = os.listdir(dir_path)

    for i in range(len(file_list)):
        file = file_list[i]
        file_path = os.path.join(dir_path, file)

        write_wb = Workbook()
        write_ws = write_wb.active
        print(file)

        for j in range(len(resource)):
            print('case', j+1)
            best_individual, best_fitness = geneticAlgorithm(resource[j], performance[j], file_path)
            write_ws.cell(j+1, 1, str(best_individual))
            write_ws.cell(j+1, 2, str(best_fitness))
            print(best_fitness)
        print()
        write_wb.save('./yoonho/result/' + file[:-5] + '.xlsx')

if __name__ == "__main__":
    main()